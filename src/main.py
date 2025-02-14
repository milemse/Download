import json
import re

import openpyxl
import psycopg2

# Leemos el archivo de configuracion
file_paths = open('/Users/fmontoya/Gasu/paths/path.json')
config = json.load(file_paths)
file_paths.close()

file_path = config['path']
period_id = config['period_id']

# Leemos el linker
file_linker = open('/Users/fmontoya/Gasu/Descarga/src/utils/linker_validation.json')
linker = json.load(file_linker)
file_linker.close()

wb = openpyxl.load_workbook(file_path)
sheet_names = wb.sheetnames
sheet_idx = 0

# Definimos regex
bbva_regex = re.compile(r'^BBVA*')
banorte_regex = re.compile(r'^BANORTE*')


# Creamos la conexion a la base de datos
conn = psycopg2.connect(dbname='provee', user='fmontoya', host='localhost', password='NZSCx81!')
# conn = psycopg2.connect(dbname='provee', user='postgres', host='localhost', password='01001Gasu')

# Creamos la consulta de lectura de consumo
query_validated_payments = f"select cd.block, cl.client as client, py.reference as payment_reference, py.done_at, tx.tax_id from main.condominium cd join main.building bd on cd.condominium_id = bd.condominium_id join main.client cl on cl.building_id = bd.building_id left join main.tax tx on cl.client_id = tx.client_id join main.account acc on cl.client_id = acc.client_id join main.payment py on acc.account_id = py.account_id where validated = true and to_download = true and py.done_at >= (select initial from main.period where period_id = {period_id}) and py.done_at <= (select final from main.period where period_id = {period_id})"

cursor = conn.cursor()
cursor.execute(query_validated_payments)
validated_payments_data = cursor.fetchall()
cursor.close()

validated_payments = []

for payment in validated_payments_data:
    validated_payments.append({
        'block': payment[0],
        'client': payment[1],
        'validation': 'OK',
        'reference': payment[2],
        'done_at': payment[3],
        'tax_id': payment[4]
    })

for sheet_name in sheet_names:
    if bbva_regex.match(sheet_name) or banorte_regex.match(sheet_name):

        for validated_payment in validated_payments:
            row_reference = validated_payment['reference'].split('-').pop()
            row_reference = row_reference.split(':')

            if sheet_idx == int(row_reference[0]):
                sheet = wb[sheet_name]
                sheet.cell(row=int(row_reference[1]), column=linker['client']['column'][sheet_idx]).value = validated_payment['client']
                sheet.cell(row=int(row_reference[1]), column=linker['block']['column'][sheet_idx]).value = validated_payment['block']
                sheet.cell(row=int(row_reference[1]), column=linker['validation']['column'][sheet_idx]).value = validated_payment['validation']
                sheet.cell(row=int(row_reference[1]), column=linker['validation_date']['column'][sheet_idx]).value = validated_payment['done_at']
                sheet.cell(row=int(row_reference[1]), column=linker['tax']['column'][sheet_idx]).value = 'R CF' if validated_payment['tax_id'] else 'R SF'

        sheet_idx += 1

file_path = file_path.replace('.xlsx', '_VALIDADO.xlsx')
wb.save(file_path)

wb.close()
conn.close()